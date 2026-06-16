from flask import Blueprint, render_template, request, redirect, url_for, g, jsonify, flash, session, send_file
from ..auth import login_required
import json, io

bp = Blueprint('reports', __name__)

# ── Data source resolvers ─────────────────────────────────────────────────────
def resolve_source(source, filters, db):
    f = filters
    if source == 'employees':
        q = """SELECT e.employee_no, e.last_name, e.first_name, e.position_title,
                      d.name as dept_name, e.payroll_group, e.daily_rate,
                      e.tax_type, e.date_hired, e.status, e.sss_no, e.philhealth_no,
                      e.pagibig_no, e.tin, e.employment_type, e.payment_method
               FROM employees e LEFT JOIN departments d ON e.department_id=d.id WHERE 1=1"""
        params = []
        if f.get('status'): q += " AND e.status=?"; params.append(f['status'])
        if f.get('payroll_group'): q += " AND e.payroll_group=?"; params.append(f['payroll_group'])
        if f.get('dept_name'): q += " AND d.name LIKE ?"; params.append(f'%{f["dept_name"]}%')
        q += " ORDER BY e.last_name, e.first_name"
        return db.execute(q, params).fetchall()

    if source == 'attendance_summary':
        df = f.get('date_from', '2026-01-01')
        dt = f.get('date_to',   '2026-12-31')
        return db.execute("""
            SELECT e.employee_no, e.last_name, e.first_name,
                   COUNT(a.id) as days_present,
                   COALESCE(SUM(a.late_minutes),0) as total_late,
                   COALESCE(SUM(a.ot_hours),0) as total_ot,
                   COALESCE(SUM(a.nd_hours),0) as total_nd,
                   COALESCE(SUM(a.total_hours),0) as total_hours,
                   SUM(CASE WHEN a.is_absent=1 THEN 1 ELSE 0 END) as days_absent
            FROM employees e LEFT JOIN attendance a
                ON a.employee_id=e.id AND a.work_date BETWEEN ? AND ?
            WHERE e.status='ACTIVE' GROUP BY e.id ORDER BY e.last_name
        """, (df, dt)).fetchall()

    if source == 'timesheet_detail':
        df = f.get('date_from', '2026-01-01')
        dt = f.get('date_to',   '2026-12-31')
        q = """SELECT e.employee_no, e.last_name, e.first_name,
                      a.work_date, a.time_in, a.time_out, a.total_hours,
                      a.ot_hours, a.late_minutes, a.is_absent, a.source,
                      a.adjustment_reason
               FROM attendance a JOIN employees e ON a.employee_id=e.id
               WHERE a.work_date BETWEEN ? AND ?"""
        params = [df, dt]
        if f.get('employee_no'):
            q += " AND e.employee_no=?"; params.append(f['employee_no'])
        q += " ORDER BY e.last_name, a.work_date"
        return db.execute(q, params).fetchall()

    if source == 'payroll_register':
        q = """SELECT e.employee_no, e.last_name, e.first_name,
                      pp.period_label, pr.basic_salary, pr.ot_amount,
                      pr.allowance_amount, pr.gross_salary, pr.sss_employee,
                      pr.philhealth_employee, pr.pagibig_employee, pr.withholding_tax,
                      pr.total_deductions, pr.net_pay, pr.status
               FROM payroll pr
               JOIN employees e ON pr.employee_id=e.id
               JOIN pay_periods pp ON pr.pay_period_id=pp.id WHERE 1=1"""
        params = []
        if f.get('period_label'): q += " AND pp.period_label LIKE ?"; params.append(f'%{f["period_label"]}%')
        q += " ORDER BY e.last_name"
        return db.execute(q, params).fetchall()

    if source == 'gov_contributions':
        return db.execute("""
            SELECT e.employee_no, e.last_name, e.first_name, e.sss_no, e.philhealth_no, e.pagibig_no,
                   SUM(pr.sss_employee) as total_sss_ee, SUM(pr.sss_employer) as total_sss_er,
                   SUM(pr.philhealth_employee) as total_phic_ee,
                   SUM(pr.pagibig_employee) as total_hdmf_ee,
                   SUM(pr.withholding_tax) as total_wtax
            FROM payroll pr JOIN employees e ON pr.employee_id=e.id
            GROUP BY e.id ORDER BY e.last_name
        """).fetchall()

    if source == 'leave_credits':
        q = """SELECT e.employee_no, e.last_name, e.first_name, lc.leave_type,
                      lc.year, lc.allocated_days, lc.used_days, lc.balance_days
               FROM leave_credits lc JOIN employees e ON lc.employee_id=e.id WHERE 1=1"""
        params = []
        if f.get('leave_type'): q += " AND lc.leave_type=?"; params.append(f['leave_type'])
        if f.get('year'): q += " AND lc.year=?"; params.append(f['year'])
        q += " ORDER BY e.last_name, lc.leave_type"
        return db.execute(q, params).fetchall()

    if source == 'loans_summary':
        q = """SELECT e.employee_no, e.last_name, e.first_name,
                      el.loan_category, el.loan_type, el.principal_amount,
                      el.outstanding_balance, el.monthly_amortization,
                      el.total_paid, el.status, el.start_date
               FROM employee_loans el JOIN employees e ON el.employee_id=e.id WHERE 1=1"""
        params = []
        if f.get('loan_category'): q += " AND el.loan_category=?"; params.append(f['loan_category'])
        if f.get('status'): q += " AND el.status=?"; params.append(f['status'])
        q += " ORDER BY e.last_name, el.loan_category"
        return db.execute(q, params).fetchall()

    return []

@bp.route('/')
@login_required
def index():
    reports = g.db.execute("""SELECT * FROM report_definitions
        ORDER BY report_module, report_name""").fetchall()
    by_module = {}
    for r in reports:
        m = r['report_module']
        if m not in by_module: by_module[m] = []
        by_module[m].append(r)
    return render_template('reports/index.html', by_module=by_module)

@bp.route('/run/<int:report_id>', methods=['GET', 'POST'])
@login_required
def run_report(report_id):
    report = g.db.execute("SELECT * FROM report_definitions WHERE id=?", (report_id,)).fetchone()
    if not report:
        flash('Report not found.', 'error')
        return redirect(url_for('reports.index'))

    columns = json.loads(report['columns_config'])
    filters_def = json.loads(report['filters_config'])
    query_cfg = json.loads(report['query_config'])
    source = query_cfg.get('source', '')

    filter_values = {}
    rows = []
    if request.method == 'POST':
        filter_values = {fd['field']: request.form.get(fd['field'], '') for fd in filters_def}
        rows = resolve_source(source, filter_values, g.db)
        g.db.execute("""INSERT INTO report_runs (report_id,run_by,parameters,row_count)
            VALUES (?,?,?,?)""",
            (report_id, session['user']['id'], json.dumps(filter_values), len(rows)))
        g.db.commit()

    return render_template('reports/run.html', report=report,
                           columns=columns, filters_def=filters_def,
                           filter_values=filter_values, rows=rows,
                           chart_type=report['chart_type'])

@bp.route('/export/<int:report_id>', methods=['POST'])
@login_required
def export_report(report_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    report = g.db.execute("SELECT * FROM report_definitions WHERE id=?", (report_id,)).fetchone()
    if not report:
        return jsonify({'error': 'Not found'}), 404
    columns = json.loads(report['columns_config'])
    filters_def = json.loads(report['filters_config'])
    query_cfg = json.loads(report['query_config'])
    filter_values = {fd['field']: request.form.get(fd['field'], '') for fd in filters_def}
    rows = resolve_source(query_cfg.get('source', ''), filter_values, g.db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report['report_name'][:31]
    ws.append([report['report_name']])
    ws['A1'].font = Font(bold=True, size=13)
    from datetime import datetime as _dt
    ws.append(['Generated: ' + _dt.now().strftime('%Y-%m-%d %H:%M')])
    ws.append([])
    hfill = PatternFill('solid', start_color='1F497D')
    hfont = Font(bold=True, color='FFFFFF')
    header_row = [col['label'] for col in columns]
    ws.append(header_row)
    for col_idx, cell in enumerate(ws[ws.max_row], 1):
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal='center')
    for row in rows:
        row_data = []
        for col in columns:
            field = col['field']
            try:
                val = row[field] if hasattr(row, '__getitem__') else getattr(row, field, '')
            except:
                val = ''
            row_data.append(val)
        ws.append(row_data)
        if col.get('format') == 'currency':
            for cell in ws[ws.max_row]:
                cell.number_format = '#,##0.00'
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    safe_name = report['report_name'].replace(' ', '_').replace('/', '-')
    return send_file(buf, as_attachment=True, download_name=f"{safe_name}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@bp.route('/create', methods=['GET', 'POST'])
@login_required
def create():
    if request.method == 'POST':
        f = request.form
        g.db.execute("""INSERT INTO report_definitions
            (report_name,report_module,description,query_config,columns_config,filters_config,chart_type,is_public,created_by)
            VALUES (?,?,?,?,?,?,?,?,?)""",
            (f['report_name'], f['report_module'], f.get('description',''),
             json.dumps({'source': f['data_source']}),
             f.get('columns_config', '[]'), f.get('filters_config', '[]'),
             f.get('chart_type','none'),
             int(f.get('is_public', 0)), session['user']['id']))
        g.db.commit()
        flash('Report created.', 'success')
        return redirect(url_for('reports.index'))
    sources = ['employees','attendance_summary','timesheet_detail','payroll_register',
               'gov_contributions','leave_credits','loans_summary']
    return render_template('reports/create.html', sources=sources)

@bp.route('/alphalist')
@login_required
def alphalist():
    records = g.db.execute("""SELECT e.last_name, e.first_name, e.middle_name, e.tin,
        SUM(pr.gross_salary) as annual_gross, SUM(pr.withholding_tax) as annual_wtax
        FROM payroll pr JOIN employees e ON pr.employee_id=e.id
        GROUP BY e.id ORDER BY e.last_name""").fetchall()
    return render_template('reports/alphalist.html', records=records)

@bp.route('/api/run/<int:report_id>', methods=['POST'])
@login_required
def api_run(report_id):
    report = g.db.execute("SELECT * FROM report_definitions WHERE id=?", (report_id,)).fetchone()
    if not report:
        return jsonify({'error': 'Not found'}), 404
    query_cfg = json.loads(report['query_config'])
    filters_def = json.loads(report['filters_config'])
    filter_values = {fd['field']: request.json.get(fd['field'], '') for fd in filters_def}
    rows = resolve_source(query_cfg.get('source', ''), filter_values, g.db)
    return jsonify({'rows': [dict(r) for r in rows], 'count': len(rows)})
