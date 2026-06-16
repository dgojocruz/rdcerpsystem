from flask import Blueprint, render_template, request, redirect, url_for, g, jsonify, flash, current_app, send_file
from ..auth import login_required
import os, io, json

bp = Blueprint('employees', __name__)

def dept_name_to_id(db, dept_name):
    """Match a position/dept string to the closest department."""
    mapping = {
        'ACCOUNTING': 'ACCTG', 'ACCTG': 'ACCTG',
        'PURCHASING': 'PURCH', 'PURCHASE': 'PURCH',
        'SALES': 'SALES', 'MARKETING': 'SALES',
        'ADMIN': 'ADMIN', 'ADMINISTRATION': 'ADMIN',
        'HR': 'HR', 'HUMAN RESOURCES': 'HR',
        'MAINTENANCE': 'MAINT', 'MAINT': 'MAINT',
        'COMPOUNDING': 'COMP', 'COMP': 'COMP',
        'VULCANIZING': 'VULC', 'VULC': 'VULC',
        'TRIMMING': 'TRIM', 'TRIM': 'TRIM',
        'QA': 'QA', 'LABORATORY': 'QA', 'QUALITY': 'QA',
        'WAREHOUSE': 'WH', 'WH': 'WH',
        'LOGISTICS': 'LOG', 'DELIVERY': 'LOG',
    }
    if not dept_name:
        return None
    upper = dept_name.upper()
    code = None
    for keyword, dept_code in mapping.items():
        if keyword in upper:
            code = dept_code
            break
    if code:
        row = db.execute("SELECT id FROM departments WHERE code=?", (code,)).fetchone()
        if row:
            return row['id']
    # fallback: fuzzy search
    for dept in db.execute("SELECT id, name, code FROM departments").fetchall():
        if dept['code'] in upper or dept['name'].upper() in upper:
            return dept['id']
    return None

@bp.route('/')
@login_required
def index():
    dept_filter = request.args.get('dept', '')
    group_filter = request.args.get('group', '')
    status_filter = request.args.get('status', 'ACTIVE')
    search = request.args.get('q', '')
    query = """SELECT e.*, d.name as dept_name, d.code as dept_code
               FROM employees e LEFT JOIN departments d ON e.department_id=d.id
               WHERE 1=1"""
    params = []
    if status_filter:
        query += " AND e.status=?"; params.append(status_filter)
    if dept_filter:
        query += " AND e.department_id=?"; params.append(dept_filter)
    if group_filter:
        query += " AND e.payroll_group=?"; params.append(group_filter)
    if search:
        query += " AND (e.last_name LIKE ? OR e.first_name LIKE ? OR e.employee_no LIKE ? OR e.position_title LIKE ?)"
        s = f'%{search}%'; params += [s, s, s, s]
    query += " ORDER BY e.last_name, e.first_name"
    employees = g.db.execute(query, params).fetchall()
    departments = g.db.execute("SELECT * FROM departments WHERE is_active=1 ORDER BY name").fetchall()
    return render_template('employees/index.html', employees=employees, departments=departments,
                           dept_filter=dept_filter, group_filter=group_filter,
                           search=search, status_filter=status_filter)

@bp.route('/add', methods=['GET','POST'])
@login_required
def add():
    if request.method == 'POST':
        f = request.form
        db = g.db
        last_id = db.execute("SELECT MAX(CAST(REPLACE(employee_no,'EMP','') AS INT)) FROM employees").fetchone()[0] or 0
        emp_no = f'EMP{str(last_id+1).zfill(4)}'
        daily = float(f.get('daily_rate') or 0)
        db.execute("""INSERT INTO employees
            (employee_no,last_name,first_name,middle_name,department_id,position_title,
             dept_type,payroll_group,payment_method,daily_rate,hourly_rate,monthly_rate,
             tax_type,tin,sss_no,philhealth_no,pagibig_no,date_hired,employment_type,
             status,gender,civil_status,phone,email,allowance_amount,bank_name,bank_account_no,biometric_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'ACTIVE',?,?,?,?,?,?,?,?)""",
            (emp_no, f['last_name'], f['first_name'], f.get('middle_name',''),
             f.get('department_id') or None, f.get('position_title',''),
             f.get('dept_type','IND'), f.get('payroll_group','MONTHLY'), f.get('payment_method','ATM'),
             daily, daily/8, float(f.get('monthly_rate') or 0), f.get('tax_type','AWE'),
             f.get('tin',''), f.get('sss_no',''), f.get('philhealth_no',''), f.get('pagibig_no',''),
             f.get('date_hired',''), f.get('employment_type','REGULAR'),
             f.get('gender',''), f.get('civil_status',''), f.get('phone',''), f.get('email',''),
             float(f.get('allowance_amount') or 0), f.get('bank_name',''),
             f.get('bank_account_no',''), f.get('biometric_id','')))
        db.commit()
        emp = db.execute("SELECT id FROM employees WHERE employee_no=?", (emp_no,)).fetchone()
        for lt, days in [('VL',5),('SL',5),('EL',3)]:
            db.execute("""INSERT OR IGNORE INTO leave_credits
                (employee_id,year,leave_type,allocated_days,used_days,balance_days)
                VALUES (?,strftime('%Y','now'),?,?,0,?)""", (emp['id'], lt, days, days))
        db.commit()
        flash(f'Employee {emp_no} added successfully.', 'success')
        return redirect(url_for('employees.index'))
    departments = g.db.execute("SELECT * FROM departments WHERE is_active=1 ORDER BY name").fetchall()
    managers = g.db.execute("SELECT id,employee_no,last_name,first_name FROM employees WHERE status='ACTIVE' ORDER BY last_name").fetchall()
    return render_template('employees/form.html', employee=None, departments=departments, managers=managers, action='Add')

@bp.route('/<int:emp_id>')
@login_required
def view(emp_id):
    emp = g.db.execute("""SELECT e.*, d.name as dept_name,
        m.last_name as mgr_last, m.first_name as mgr_first
        FROM employees e LEFT JOIN departments d ON e.department_id=d.id
        LEFT JOIN employees m ON e.manager_id=m.id
        WHERE e.id=?""", (emp_id,)).fetchone()
    if not emp: return redirect(url_for('employees.index'))
    loans = g.db.execute("SELECT * FROM employee_loans WHERE employee_id=? AND status='ACTIVE'", (emp_id,)).fetchall()
    leaves = g.db.execute("SELECT * FROM leave_credits WHERE employee_id=? AND year=strftime('%Y','now')", (emp_id,)).fetchall()
    recent_payroll = g.db.execute("""SELECT pr.*, pp.period_label FROM payroll pr
        JOIN pay_periods pp ON pr.pay_period_id=pp.id
        WHERE pr.employee_id=? ORDER BY pp.date_from DESC LIMIT 6""", (emp_id,)).fetchall()

    # Timekeeping history
    month_filter = request.args.get('tk_month', '')
    tk_query = """SELECT work_date, time_in, time_out, total_hours, regular_hours,
        ot_hours, nd_hours, late_minutes, undertime_minutes,
        is_absent, is_holiday, holiday_type, is_rest_day, source, remarks
        FROM attendance WHERE employee_id=?"""
    tk_params = [emp_id]
    if month_filter:
        tk_query += " AND strftime('%Y-%m', work_date)=?"
        tk_params.append(month_filter)
    tk_query += " ORDER BY work_date DESC LIMIT 60"
    attendance_history = g.db.execute(tk_query, tk_params).fetchall()

    # Attendance summary stats
    tk_stats = g.db.execute("""SELECT
        COUNT(*) as total_days,
        SUM(CASE WHEN is_absent=0 AND is_rest_day=0 AND is_holiday=0 THEN 1 ELSE 0 END) as present_days,
        SUM(CASE WHEN is_absent=1 THEN 1 ELSE 0 END) as absent_days,
        SUM(CASE WHEN late_minutes>0 THEN 1 ELSE 0 END) as late_days,
        COALESCE(SUM(ot_hours),0) as total_ot,
        COALESCE(SUM(late_minutes),0) as total_late_min,
        COALESCE(SUM(total_hours),0) as total_hours_worked
        FROM attendance WHERE employee_id=?
        AND (? = '' OR strftime('%Y-%m', work_date)=?)
    """, (emp_id, month_filter, month_filter)).fetchone()

    # Schedule history
    sch_month = request.args.get('sch_month', '')
    sch_query = """SELECT es.schedule_date, es.is_rest_day, es.schedule_type,
        sd.shift_name, sd.time_in, sd.time_out, sd.break_minutes, sd.color_hex
        FROM employee_schedules es
        LEFT JOIN shift_definitions sd ON es.shift_id=sd.id
        WHERE es.employee_id=?"""
    sch_params = [emp_id]
    if sch_month:
        sch_query += " AND strftime('%Y-%m', es.schedule_date)=?"
        sch_params.append(sch_month)
    sch_query += " ORDER BY es.schedule_date DESC LIMIT 60"
    schedule_history = g.db.execute(sch_query, sch_params).fetchall()

    departments = g.db.execute("SELECT * FROM departments WHERE is_active=1").fetchall()
    return render_template('employees/view.html', emp=emp, loans=loans, leaves=leaves,
                           recent_payroll=recent_payroll, departments=departments,
                           attendance_history=attendance_history, tk_stats=tk_stats,
                           tk_month=month_filter, schedule_history=schedule_history,
                           sch_month=sch_month)

@bp.route('/<int:emp_id>/edit', methods=['GET','POST'])
@login_required
def edit(emp_id):
    emp = g.db.execute("SELECT * FROM employees WHERE id=?", (emp_id,)).fetchone()
    if not emp: return redirect(url_for('employees.index'))
    if request.method == 'POST':
        f = request.form
        daily = float(f.get('daily_rate') or emp['daily_rate'])
        g.db.execute("""UPDATE employees SET
            last_name=?,first_name=?,middle_name=?,department_id=?,position_title=?,
            dept_type=?,payroll_group=?,payment_method=?,daily_rate=?,hourly_rate=?,
            monthly_rate=?,tax_type=?,tin=?,sss_no=?,philhealth_no=?,pagibig_no=?,
            date_hired=?,employment_type=?,gender=?,civil_status=?,phone=?,email=?,
            allowance_amount=?,bank_name=?,bank_account_no=?,biometric_id=?,
            manager_id=?,updated_at=datetime('now') WHERE id=?""",
            (f['last_name'], f['first_name'], f.get('middle_name',''),
             f.get('department_id') or None, f.get('position_title',''),
             f.get('dept_type','IND'), f.get('payroll_group','MONTHLY'), f.get('payment_method','ATM'),
             daily, daily/8, float(f.get('monthly_rate') or 0), f.get('tax_type','AWE'),
             f.get('tin',''), f.get('sss_no',''), f.get('philhealth_no',''), f.get('pagibig_no',''),
             f.get('date_hired',''), f.get('employment_type','REGULAR'),
             f.get('gender',''), f.get('civil_status',''), f.get('phone',''), f.get('email',''),
             float(f.get('allowance_amount') or 0), f.get('bank_name',''),
             f.get('bank_account_no',''), f.get('biometric_id',''),
             f.get('manager_id') or None, emp_id))
        g.db.commit()
        flash('Employee updated.', 'success')
        return redirect(url_for('employees.view', emp_id=emp_id))
    departments = g.db.execute("SELECT * FROM departments WHERE is_active=1 ORDER BY name").fetchall()
    managers = g.db.execute("""SELECT id,employee_no,last_name,first_name
        FROM employees WHERE status='ACTIVE' AND id != ? ORDER BY last_name""", (emp_id,)).fetchall()
    return render_template('employees/form.html', employee=emp, departments=departments, managers=managers, action='Edit')

@bp.route('/import', methods=['GET','POST'])
@login_required
def import_excel():
    if request.method == 'GET':
        return render_template('employees/import.html')
    import openpyxl
    file = request.files.get('file')
    if not file:
        flash('No file selected.', 'error')
        return redirect(url_for('employees.import_excel'))
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        flash(f'Cannot read file: {e}', 'error')
        return redirect(url_for('employees.import_excel'))

    seen_keys = set()
    employees_to_import = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # Find the header row (contains 'Employee Last Name')
        header_idx = None
        for i, row in enumerate(rows):
            if any(str(c or '').strip() == 'Employee Last Name' for c in row):
                header_idx = i
                break
        if header_idx is None:
            continue
        for row in rows[header_idx + 1:]:
            last = str(row[1] or '').strip()
            first = str(row[2] or '').strip()
            if not last or not first or last == 'Employee Last Name':
                continue
            key = f"{last.upper()}|{first.upper()}"
            if key in seen_keys:
                continue
            seen_keys.add(key)
            try:
                daily = float(str(row[10] or 0).replace(',', ''))
            except:
                daily = 0
            employees_to_import.append({
                'last_name': last,
                'first_name': first,
                'middle_name': str(row[3] or '').strip(),
                'payroll_group': str(row[4] or 'MONTHLY').strip(),
                'tin': str(row[5] or '').strip(),
                'position_title': str(row[6] or '').strip(),
                'dept_type': str(row[7] or 'IND').strip().strip(),
                'tax_type': str(row[8] or 'AWE').strip(),
                'payment_method': str(row[9] or 'ATM').strip(),
                'daily_rate': daily,
                'hourly_rate': daily / 8 if daily else 0,
            })

    if not employees_to_import:
        flash('No employee data found in file. Check that your file has the correct column headers.', 'error')
        return redirect(url_for('employees.import_excel'))

    # Preview mode
    if request.form.get('action') == 'preview':
        return render_template('employees/import_preview.html',
                               employees=employees_to_import, count=len(employees_to_import))

    # Actual import
    db = g.db
    last_id = db.execute("SELECT MAX(CAST(REPLACE(employee_no,'EMP','') AS INT)) FROM employees").fetchone()[0] or 0
    imported = 0
    skipped = 0
    errors = []
    for emp in employees_to_import:
        exists = db.execute("""SELECT id FROM employees
            WHERE last_name=? AND first_name=?""",
            (emp['last_name'], emp['first_name'])).fetchone()
        if exists:
            skipped += 1
            continue
        last_id += 1
        emp_no = f"EMP{str(last_id).zfill(4)}"
        dept_id = dept_name_to_id(db, emp['position_title'])
        try:
            db.execute("""INSERT INTO employees
                (employee_no,last_name,first_name,middle_name,department_id,
                 position_title,dept_type,payroll_group,payment_method,
                 daily_rate,hourly_rate,tax_type,tin,status,employment_type)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,'ACTIVE','REGULAR')""",
                (emp_no, emp['last_name'], emp['first_name'], emp['middle_name'],
                 dept_id, emp['position_title'], emp['dept_type'],
                 emp['payroll_group'], emp['payment_method'],
                 emp['daily_rate'], emp['hourly_rate'],
                 emp['tax_type'], emp['tin']))
            eid = db.execute("SELECT last_insert_rowid() as id").fetchone()['id']
            for lt, days in [('VL',5),('SL',5),('EL',3)]:
                db.execute("""INSERT OR IGNORE INTO leave_credits
                    (employee_id,year,leave_type,allocated_days,used_days,balance_days)
                    VALUES (?,strftime('%Y','now'),?,?,0,?)""", (eid, lt, days, days))
            imported += 1
        except Exception as e:
            errors.append(f"{emp['last_name']}, {emp['first_name']}: {e}")
    db.commit()
    flash(f'Import complete: {imported} imported, {skipped} already existed.', 'success')
    if errors:
        flash(f'Errors: {"; ".join(errors[:3])}', 'error')
    return redirect(url_for('employees.index'))

@bp.route('/org-chart')
@login_required
def org_chart():
    employees = g.db.execute("""SELECT e.id, e.employee_no, e.last_name, e.first_name,
        e.position_title, e.manager_id, e.dept_type, e.status,
        d.name as dept_name, d.code as dept_code
        FROM employees e LEFT JOIN departments d ON e.department_id=d.id
        WHERE e.status='ACTIVE' ORDER BY e.last_name""").fetchall()
    emp_list = [dict(e) for e in employees]
    return render_template('employees/org_chart.html', employees_json=emp_list)

@bp.route('/api/list')
@login_required
def api_list():
    emps = g.db.execute("""SELECT id,employee_no,last_name,first_name,position_title,
        payroll_group,daily_rate,tax_type,status FROM employees WHERE status='ACTIVE'
        ORDER BY last_name""").fetchall()
    return jsonify([dict(e) for e in emps])

@bp.route('/download-template')
@login_required
def download_template():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Employees'
    headers = ['Employee Last Name','Employee First Name','Employee Middle Name',
               'Payroll Group','TAX IDENTIFICATION NO.','Department/Position',
               'Dept. Type','Tax Type','Payment Method','Daily Rate','SSS No',
               'PhilHealth No','Pag-IBIG No','Bank Name','Bank Account No',
               'Date Hired','Gender','Civil Status']
    ws.append(headers)
    ws.append(['DELA CRUZ','JUAN','SANTOS','MONTHLY','123-456-789',
               'Accounting','ADM','AWE','ATM','888.22','34-123456-7',
               '12-123456789-0','1234-5678-9012','BDO','123456789','2024-01-15','Male','Single'])
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='employee_import_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
