from flask import Blueprint, render_template, request, redirect, url_for, g, jsonify, flash, current_app, send_file
from ..auth import login_required
from datetime import datetime
import io

bp = Blueprint('payroll', __name__)

def get_cfg(db):
    rows = db.execute("SELECT config_key, config_value FROM payroll_config").fetchall()
    return {r['config_key']: float(r['config_value']) if r['config_value'].replace('.','').isdigit() else r['config_value'] for r in rows}

def compute_sss(monthly, cfg):
    msc_table = [(0,4249,4000),(4250,4749,4500),(4750,5249,5000),(5250,5749,5500),
                 (5750,6249,6000),(6250,6749,6500),(6750,7249,7000),(7250,7749,7500),
                 (7750,8249,8000),(8250,8749,8500),(8750,9249,9000),(9250,9749,9500),
                 (9750,10249,10000),(10250,10749,10500),(10750,11249,11000),(11250,11749,11500),
                 (11750,12249,12000),(12250,12749,12500),(12750,13249,13000),(13250,13749,13500),
                 (13750,14249,14000),(14250,14749,14500),(14750,15249,15000),(15250,15749,15500),
                 (15750,16249,16000),(16250,16749,16500),(16750,17249,17000),(17250,17749,17500),
                 (17750,18249,18000),(18250,18749,18500),(18750,19249,19000),(19250,19749,19500),
                 (19750,20249,20000),(20250,999999,20250)]
    msc = 20250
    for lo, hi, credit in msc_table:
        if lo <= monthly <= hi:
            msc = credit; break
    ee = round(msc * cfg.get('SSS_EE_RATE', 0.045), 2)
    er = round(msc * cfg.get('SSS_ER_RATE', 0.085), 2)
    wisp = round(max(msc - 20000, 0) * 0.01, 2)
    return ee, er, wisp

def compute_philhealth(monthly, cfg):
    rate = cfg.get('PHIC_RATE', 0.05)
    max_sal = cfg.get('PHIC_MAX_SAL', 100000)
    min_con = cfg.get('PHIC_MIN_CON', 500)
    basis = min(monthly, max_sal)
    total = max(basis * rate, min_con)
    return round(total / 2, 2), round(total / 2, 2)

def compute_pagibig(monthly, cfg):
    max_sal = cfg.get('HDMF_MAX_SAL', 10000)
    max_con = cfg.get('HDMF_MAX_CON', 200)
    ee_rate = 0.01 if monthly <= 1500 else cfg.get('HDMF_EE_RATE', 0.02)
    er_rate = cfg.get('HDMF_ER_RATE', 0.02)
    basis = min(monthly, max_sal)
    return min(round(basis * ee_rate, 2), max_con), min(round(basis * er_rate, 2), max_con)

def compute_withholding_tax(taxable_monthly, tax_type='AWE'):
    if tax_type == 'MWE':
        return 0
    brackets = [(0,20833,0,0),(20833,33332,0,0.15),(33332,66666,1875,0.20),
                (66666,166666,8541.80,0.25),(166666,666666,33541.80,0.30),(666666,1e9,183541.80,0.35)]
    annual = taxable_monthly * 12
    tax = 0
    for lo, hi, base, rate in brackets:
        if annual > lo:
            tax = base + (min(annual, hi) - lo) * rate
    return max(round(tax / 12, 2), 0)

@bp.route('/')
@login_required
def index():
    periods = g.db.execute("SELECT * FROM pay_periods ORDER BY date_from DESC LIMIT 20").fetchall()
    return render_template('payroll/index.html', periods=periods)

@bp.route('/period/new', methods=['GET','POST'])
@login_required
def new_period():
    if request.method == 'POST':
        f = request.form
        g.db.execute("""INSERT INTO pay_periods
            (period_type,period_label,date_from,date_to,payroll_group,status)
            VALUES (?,?,?,?,?,'OPEN')""",
            (f['period_type'],f['period_label'],f['date_from'],f['date_to'],f.get('payroll_group','ALL')))
        g.db.commit()
        pid = g.db.execute("SELECT last_insert_rowid() as id").fetchone()['id']
        return redirect(url_for('payroll.compute', period_id=pid))
    return render_template('payroll/new_period.html')

@bp.route('/period/<int:period_id>/compute', methods=['GET','POST'])
@login_required
def compute(period_id):
    period = g.db.execute("SELECT * FROM pay_periods WHERE id=?", (period_id,)).fetchone()
    if not period: return redirect(url_for('payroll.index'))
    if request.method == 'POST':
        cfg = get_cfg(g.db)
        _do_compute(period_id, period, cfg, g.db)
        flash('Payroll computed successfully.', 'success')
        return redirect(url_for('payroll.view_period', period_id=period_id))
    employees = g.db.execute("""SELECT e.* FROM employees e WHERE e.status='ACTIVE'
        AND (? = 'ALL' OR e.payroll_group = ?)""",
        (period['payroll_group'], period['payroll_group'])).fetchall()
    return render_template('payroll/compute.html', period=period, employees=employees)

def _do_compute(period_id, period, cfg, db):
    employees = db.execute("""SELECT e.* FROM employees e WHERE e.status='ACTIVE'
        AND (? = 'ALL' OR e.payroll_group = ?)""",
        (period['payroll_group'], period['payroll_group'])).fetchall()
    for emp in employees:
        att = db.execute("""SELECT
            COALESCE(SUM(total_hours),0) as th, COALESCE(SUM(ot_hours),0) as ot,
            COALESCE(SUM(nd_hours),0) as nd, COALESCE(SUM(late_minutes),0) as late,
            COALESCE(SUM(undertime_minutes),0) as ut
            FROM attendance WHERE employee_id=? AND work_date BETWEEN ? AND ?""",
            (emp['id'], period['date_from'], period['date_to'])).fetchone()
        daily = emp['daily_rate'] or cfg.get('NCR_MIN_WAGE', 645)
        hr = daily / 8
        rpm = hr / 60
        reg_amt = (att['th'] - att['ot']) * hr
        late_amt = att['late'] * rpm
        ut_amt = att['ut'] * rpm
        basic = max(reg_amt - late_amt - ut_amt, 0)
        ot_rate = cfg.get('OT_RATE', 1.25)
        ot_amt = att['ot'] * hr * ot_rate
        nd_rate = cfg.get('ND_RATE', 0.10)
        nd_amt = att['nd'] * hr * nd_rate
        allowance = emp['allowance_amount'] or 0
        gross = basic + ot_amt + nd_amt + allowance
        monthly_equiv = daily * 26
        sss_ee, sss_er, wisp = compute_sss(monthly_equiv, cfg)
        ph_ee, ph_er = compute_philhealth(monthly_equiv, cfg)
        pi_ee, pi_er = compute_pagibig(monthly_equiv, cfg)
        taxable = gross - sss_ee - ph_ee - pi_ee
        wtax = compute_withholding_tax(taxable, emp['tax_type'])
        loans = db.execute("""SELECT loan_category, SUM(monthly_amortization) as amt
            FROM employee_loans WHERE employee_id=? AND status='ACTIVE'
            GROUP BY loan_category""", (emp['id'],)).fetchall()
        sss_loan = sum(l['amt'] for l in loans if l['loan_category']=='SSS')
        pi_loan = sum(l['amt'] for l in loans if l['loan_category']=='PAGIBIG')
        pers_loan = sum(l['amt'] for l in loans if l['loan_category'] in ('PERSONAL','OTHER'))
        total_ded = sss_ee + wisp + ph_ee + pi_ee + wtax + sss_loan + pi_loan + pers_loan
        net = max(gross - total_ded, 0)
        # Hours paid = regular hours actually compensated; unpaid = late+undertime in hours
        hours_paid   = round((att['th'] - att['ot']), 2)
        unpaid_hours = round((att['late'] + att['ut']) / 60, 2)

        db.execute("""INSERT OR REPLACE INTO payroll
            (pay_period_id,employee_id,payroll_group,total_hours,regular_amount,
             late_amount,undertime_amount,basic_salary,ot_hours,ot_amount,
             nd_hours,nd_amount,allowance_amount,gross_salary,
             sss_employee,sss_wisp,philhealth_employee,pagibig_employee,withholding_tax,
             sss_loan,pagibig_loan,personal_loan,total_deductions,net_pay,
             sss_employer,philhealth_employer,pagibig_employer,
             hours_paid,unpaid_hours,
             status,computed_at,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,
                    'COMPUTED',datetime('now'),datetime('now'))""",
            (period_id, emp['id'], emp['payroll_group'], att['th'], reg_amt,
             late_amt, ut_amt, basic, att['ot'], ot_amt, att['nd'], nd_amt,
             allowance, gross, sss_ee, wisp, ph_ee, pi_ee, wtax,
             sss_loan, pi_loan, pers_loan, total_ded, net, sss_er, ph_er, pi_er,
             hours_paid, unpaid_hours))
    db.execute("""UPDATE pay_periods SET status='COMPUTED',
        processed_by=1, processed_at=datetime('now') WHERE id=?""", (period_id,))
    db.commit()

@bp.route('/period/<int:period_id>')
@login_required
def view_period(period_id):
    period = g.db.execute("SELECT * FROM pay_periods WHERE id=?", (period_id,)).fetchone()
    records = g.db.execute("""SELECT pr.*, e.employee_no, e.last_name, e.first_name,
        e.position_title, d.name as dept_name
        FROM payroll pr JOIN employees e ON pr.employee_id=e.id
        LEFT JOIN departments d ON e.department_id=d.id
        WHERE pr.pay_period_id=? ORDER BY e.last_name""", (period_id,)).fetchall()
    totals = g.db.execute("""SELECT SUM(gross_salary) as gross, SUM(total_deductions) as ded,
        SUM(net_pay) as net, COUNT(*) as count FROM payroll WHERE pay_period_id=?""",
        (period_id,)).fetchone()
    return render_template('payroll/view_period.html', period=period, records=records, totals=totals)

@bp.route('/period/<int:period_id>/approve', methods=['POST'])
@login_required
def approve(period_id):
    g.db.execute("UPDATE pay_periods SET status='APPROVED',approved_by=1,approved_at=datetime('now') WHERE id=?", (period_id,))
    g.db.execute("UPDATE payroll SET status='APPROVED' WHERE pay_period_id=?", (period_id,))
    g.db.commit()
    flash('Payroll approved.', 'success')
    return redirect(url_for('payroll.view_period', period_id=period_id))

@bp.route('/period/<int:period_id>/release', methods=['POST'])
@login_required
def release(period_id):
    g.db.execute("UPDATE pay_periods SET status='RELEASED',released_at=datetime('now') WHERE id=?", (period_id,))
    g.db.execute("UPDATE payroll SET status='RELEASED' WHERE pay_period_id=?", (period_id,))
    g.db.commit()
    flash('Payroll released.', 'success')
    return redirect(url_for('payroll.view_period', period_id=period_id))

@bp.route('/payslip/<int:payroll_id>')
@login_required
def payslip(payroll_id):
    record = g.db.execute("""SELECT pr.*, e.*, pp.period_label, pp.date_from, pp.date_to,
        d.name as dept_name FROM payroll pr
        JOIN employees e ON pr.employee_id=e.id
        JOIN pay_periods pp ON pr.pay_period_id=pp.id
        LEFT JOIN departments d ON e.department_id=d.id
        WHERE pr.id=?""", (payroll_id,)).fetchone()
    company = g.db.execute("SELECT * FROM company_settings LIMIT 1").fetchone()
    return render_template('payroll/payslip.html', r=record, company=company)

@bp.route('/period/<int:period_id>/export')
@login_required
def export_excel(period_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    period = g.db.execute("SELECT * FROM pay_periods WHERE id=?", (period_id,)).fetchone()
    records = g.db.execute("""SELECT pr.*, e.employee_no, e.last_name, e.first_name,
        e.daily_rate, e.payment_method, d.name as dept_name
        FROM payroll pr JOIN employees e ON pr.employee_id=e.id
        LEFT JOIN departments d ON e.department_id=d.id
        WHERE pr.pay_period_id=? ORDER BY e.last_name""", (period_id,)).fetchall()
    company = g.db.execute("SELECT * FROM company_settings LIMIT 1").fetchone()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Payroll Register'
    ws.merge_cells('A1:S1')
    ws['A1'] = company['company_name'] if company else 'Company'
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:S2')
    ws['A2'] = f"PAY PERIOD: {period['period_label']}"
    ws['A2'].alignment = Alignment(horizontal='center')
    hfill = PatternFill('solid', start_color='1F497D')
    hfont = Font(bold=True, color='FFFFFF', size=9)
    headers = ['EMP NO','NAME','DEPT','GROUP','DAILY RATE','BASIC','OT','ND','ALLOWANCE','GROSS',
               'SSS','PHILHEALTH','PAGIBIG','W.TAX','SSS LOAN','HDMF LOAN','PERS LOAN','TOTAL DED','NET PAY']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal='center')
    for ri, rec in enumerate(records, 4):
        row = [rec['employee_no'], f"{rec['last_name']}, {rec['first_name']}",
               rec['dept_name'] or '', rec['payroll_group'], rec['daily_rate'],
               rec['basic_salary'], rec['ot_amount'], rec['nd_amount'],
               rec['allowance_amount'], rec['gross_salary'],
               rec['sss_employee'], rec['philhealth_employee'], rec['pagibig_employee'],
               rec['withholding_tax'], rec['sss_loan'], rec['pagibig_loan'],
               rec['personal_loan'], rec['total_deductions'], rec['net_pay']]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=col, value=val)
            if col >= 5: cell.number_format = '#,##0.00'
            if ri % 2 == 0: cell.fill = PatternFill('solid', start_color='EBF1F5')
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
        download_name=f"payroll_{period['period_label'].replace(' ','_')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/period/<int:period_id>/delete', methods=['POST'])
@login_required
def delete_period(period_id):
    period = g.db.execute("SELECT * FROM pay_periods WHERE id=?", (period_id,)).fetchone()
    if not period:
        flash("Pay period not found.", "error")
        return redirect(url_for("payroll.index"))
    if period["status"] == "RELEASED":
        flash("Cannot delete a released payroll period.", "error")
        return redirect(url_for("payroll.index"))
    if period["status"] == "APPROVED":
        flash("Cannot delete an approved payroll. Please contact your system administrator.", "error")
        return redirect(url_for("payroll.index"))
    label = period["period_label"]
    g.db.execute("DELETE FROM payroll WHERE pay_period_id=?", (period_id,))
    g.db.execute("DELETE FROM pay_periods WHERE id=?", (period_id,))
    g.db.commit()
    flash(f"Payroll period \"{label}\" and all its records have been deleted.", "success")
    return redirect(url_for("payroll.index"))
@bp.route('/config')
@login_required
def config():
    cfg_rows = g.db.execute("SELECT * FROM payroll_config ORDER BY config_group, config_label").fetchall()
    groups = {}
    for r in cfg_rows:
        grp = r['config_group']
        if grp not in groups: groups[grp] = []
        groups[grp].append(r)
    return render_template('payroll/config.html', groups=groups)

@bp.route('/config/update', methods=['POST'])
@login_required
def update_config():
    for key, val in request.form.items():
        if key.startswith('cfg_'):
            config_key = key[4:]
            g.db.execute("""UPDATE payroll_config SET config_value=?, updated_by=1, updated_at=datetime('now')
                           WHERE config_key=?""", (val, config_key))
    g.db.commit()
    flash('Payroll configuration updated.', 'success')
    return redirect(url_for('payroll.config'))

@bp.route('/thirteenth-month', methods=['GET','POST'])
@login_required
def thirteenth_month():
    if request.method == 'POST':
        year = int(request.form.get('year', 2025))
        group = request.form.get('payroll_group', 'ALL')
        results = _compute_13th_month(year, group, g.db)
        return render_template('payroll/thirteenth_month.html', results=results, year=year, group=group)
    from datetime import date
    return render_template('payroll/thirteenth_month.html', results=None, year=date.today().year - 1, group='ALL')

def _compute_13th_month(year, group, db):
    employees = db.execute("""SELECT e.* FROM employees e WHERE e.status IN ('ACTIVE','RESIGNED')
        AND (? = 'ALL' OR e.payroll_group = ?)""", (group, group)).fetchall()
    results = []
    for emp in employees:
        total_basic = db.execute("""SELECT COALESCE(SUM(pr.basic_salary), 0) as total
            FROM payroll pr JOIN pay_periods pp ON pr.pay_period_id = pp.id
            WHERE pr.employee_id=? AND strftime('%Y', pp.date_from)=?""",
            (emp['id'], str(year))).fetchone()['total']
        thirteenth = round(total_basic / 12, 2)
        months_worked = db.execute("""SELECT COUNT(DISTINCT strftime('%Y-%m', pp.date_from)) as months
            FROM payroll pr JOIN pay_periods pp ON pr.pay_period_id = pp.id
            WHERE pr.employee_id=? AND strftime('%Y', pp.date_from)=?""",
            (emp['id'], str(year))).fetchone()['months']
        results.append({
            'employee_no': emp['employee_no'],
            'last_name': emp['last_name'],
            'first_name': emp['first_name'],
            'payroll_group': emp['payroll_group'],
            'total_basic': total_basic,
            'months_worked': months_worked,
            'thirteenth_month': thirteenth,
            'tax_type': emp['tax_type'],
            'id': emp['id'],
        })
    return sorted(results, key=lambda x: x['last_name'])

@bp.route('/backpay/<int:emp_id>', methods=['GET','POST'])
@login_required
def backpay(emp_id):
    emp = g.db.execute("""SELECT e.*, d.name as dept_name FROM employees e
        LEFT JOIN departments d ON e.department_id=d.id WHERE e.id=?""", (emp_id,)).fetchone()
    company = g.db.execute("SELECT * FROM company_settings LIMIT 1").fetchone()
    if request.method == 'POST':
        f = request.form
        last_day = f.get('last_day_worked', '')
        separation_type = f.get('separation_type', 'RESIGNED')
        unpaid_days = float(f.get('unpaid_days', 0))
        unused_vl = float(f.get('unused_vl', 0))
        unused_sl = float(f.get('unused_sl', 0))
        pro_rated_13th = float(f.get('pro_rated_13th', 0))
        separation_pay = float(f.get('separation_pay', 0))
        # Compute totals
        daily = emp['daily_rate'] or 0
        unpaid_salary = unpaid_days * daily
        vl_cash = unused_vl * daily
        sl_cash = unused_sl * daily
        gross_backpay = unpaid_salary + vl_cash + sl_cash + pro_rated_13th + separation_pay
        # Contributions on backpay
        cfg = get_cfg(g.db)
        monthly_eq = daily * 26
        sss_ee, _, _ = compute_sss(monthly_eq, cfg)
        ph_ee, _ = compute_philhealth(monthly_eq, cfg)
        pi_ee, _ = compute_pagibig(monthly_eq, cfg)
        # Loans outstanding
        loans = g.db.execute("""SELECT loan_type, outstanding_balance
            FROM employee_loans WHERE employee_id=? AND status='ACTIVE'""", (emp_id,)).fetchall()
        total_loans = sum(l['outstanding_balance'] for l in loans)
        total_ded = sss_ee + ph_ee + pi_ee + total_loans
        net_backpay = max(gross_backpay - total_ded, 0)
        bp_data = {
            'emp': emp, 'company': company, 'last_day': last_day,
            'separation_type': separation_type,
            'unpaid_days': unpaid_days, 'unpaid_salary': unpaid_salary,
            'unused_vl': unused_vl, 'vl_cash': vl_cash,
            'unused_sl': unused_sl, 'sl_cash': sl_cash,
            'pro_rated_13th': pro_rated_13th, 'separation_pay': separation_pay,
            'gross_backpay': gross_backpay,
            'sss_ee': sss_ee, 'ph_ee': ph_ee, 'pi_ee': pi_ee,
            'loans': loans, 'total_loans': total_loans,
            'total_ded': total_ded, 'net_backpay': net_backpay,
        }
        return render_template('payroll/backpay_slip.html', bp=bp_data)
    leave_credits = g.db.execute("""SELECT leave_type, balance_days FROM leave_credits
        WHERE employee_id=? AND year=strftime('%Y','now')""", (emp_id,)).fetchall()
    leave_map = {lc['leave_type']: lc['balance_days'] for lc in leave_credits}
    return render_template('payroll/backpay_form.html', emp=emp, leave_map=leave_map)
